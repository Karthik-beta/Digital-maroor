from rest_framework import generics, filters, status
from rest_framework.views import APIView
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.pagination import PageNumberPagination
from django.core.cache import cache
from django.db.models import Q, F, Count, Sum
from django.db.models.signals import post_save, post_delete
from django.dispatch import receiver
from datetime import datetime, timedelta, date
import pytz
from django.utils.timezone import make_aware, timezone, now
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.views.generic import View
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle

from resource.models import Employee, Attendance, Logs
from . import serializers
from .services import generate_unique_ids, check_employee_id

from config.models import Company, Location

class DefaultPagination(PageNumberPagination):
    """
    Default pagination class with page size set to 10.
    """
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


class EmployeeIdGet(APIView):
    """
    API view for generating and retrieving employee IDs.
    """

    def get(self, request, *args, **kwargs):
        """
        Generate unique employee ID and device enroll ID.
        """
        employee_id, device_enroll_id = generate_unique_ids()
        return Response({"employee_id": employee_id, "device_enroll_id": device_enroll_id}, status=status.HTTP_200_OK)


# @receiver(post_save, sender=Employee)
# @receiver(post_delete, sender=Employee)
# def invalidate_employee_cache(sender, instance, **kwargs):
#     """
#     Signal handler to invalidate and reload the employee cache.
#     """
#     # Invalidate employee cache
#     cache.delete('employees')
#     cache.set('employees', Employee.objects.order_by('-id').all())

# @receiver(post_save, sender=Employee)
# def reload_employee_cache(sender, instance, created, **kwargs):
#     """
#     Signal handler to reload the employee cache when a new record is added or updated.
#     """
#     # Reload employee cache only if a new record is created or an existing record is updated
#     if created or not instance._state.adding:
#         cache.set('employees', Employee.objects.order_by('-id').all())


# # Pre-load data into cache
# cache.get_or_set('employees', Employee.objects.order_by('-id').all(), timeout=3600)


class EmployeeListCreate(generics.ListCreateAPIView):
    """
    API view for listing and creating employees.
    """
    queryset = Employee.objects.all()
    serializer_class = serializers.EmployeeSerializer
    pagination_class = DefaultPagination
    filter_backends = (DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter)
    filter_fields = '__all__'
    search_fields = ['employee_id', 'employee_name', 'job_status']
    ordering_fields = '__all__'
    ordering = ['-id']

    # def get_queryset(self):
    #     """
    #     Get the cached queryset for listing employees with optional search query.
    #     """
    #     queryset = cache.get('employees')
    #     if queryset is None:
    #         print("Fetching queryset from the database.")
    #         queryset = Employee.objects.order_by('-id').all()
    #         cache.set('employees', queryset)
    #     else:
    #         print("Fetching queryset from the cache.")
    #     return queryset

    def search_queryset(self, queryset, search_query):
        """
        Filter the queryset based on the search query.
        """
        return queryset.filter(
            Q(employee_id__icontains=search_query) |
            Q(employee_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            # Q(phone_no__icontains=search_query) |
            # Q(designation_name__icontains=search_query) |
            # Q(department_name__icontains=search_query) |
            # Q(location_name__icontains=search_query) |
            Q(job_status__icontains=search_query)
        )

    def get(self, request, *args, **kwargs):
        """
        Get the list of employees with optional search query.
        """
        search_query = self.request.GET.get('search')
        queryset = self.get_queryset()

        if search_query:
            queryset = self.search_queryset(queryset, search_query)

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class EmployeeRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    """
    API view for retrieving, updating, and deleting an employee.
    """
    queryset = Employee.objects.all()
    serializer_class = serializers.EmployeeSerializer
    lookup_url_kwarg = "id"


class AttendanceListCreate(generics.ListCreateAPIView):
    """
    API view for listing and creating attendance records.
    """
    # queryset = Attendance.objects.all()
    serializer_class = serializers.AttendanceSerializer
    pagination_class = DefaultPagination
    filter_backends = (DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter)
    filter_fields = '__all__'
    search_fields = ['employee_id', 'employee_name', 'device_enroll_id', 'logdate', 'shift', 
                    '=shift_status', 'first_logtime', 'last_logtime', 'total_time', 'late_entry', 
                    'early_exit', 'overtime', 'company_name', 'location_name']
    ordering_fields = '__all__'
    ordering = ['-logdate']

    def get_queryset(self):
        """
        Get the queryset for listing attendance records with optional search query.
        """
        search_query = self.request.GET.get('search')
        date_query = self.request.GET.get('date')

        # Get date range parameters
        date_from = self.request.GET.get('date_from') 
        date_to = self.request.GET.get('date_to') 

        late_entry_not_null = self.request.GET.get('late_entry')
        early_exit_not_null = self.request.GET.get('early_exit')
        overtime_not_null = self.request.GET.get('overtime')
        employee_id = self.request.GET.get('employeeid')  
        company_names = self.request.GET.get('company_name')
        location_names = self.request.GET.get('location_name') 
        department_names = self.request.GET.get('department_name')
        designation_names = self.request.GET.get('designation_name')

        queryset = Attendance.objects.order_by('-logdate').all()

        if search_query:
            queryset = queryset.filter(
                Q(employeeid__employee_name__icontains=search_query) |
                Q(employeeid__device_enroll_id__icontains=search_query) |
                Q(logdate__icontains=search_query) |
                Q(shift_status__iexact=search_query) |
                Q(first_logtime__icontains=search_query) |
                Q(last_logtime__icontains=search_query) |
                Q(total_time__icontains=search_query) |
                Q(late_entry__icontains=search_query) |
                Q(employeeid__company__name__icontains=search_query) |
                Q(employeeid__location__name__icontains=search_query) |
                Q(employeeid__department__name__icontains=search_query) |
                Q(employeeid__designation__name__icontains=search_query)
            )

        if date_query:
            try:
                # Convert date_query to a datetime object
                date_obj = datetime.strptime(date_query, '%m-%d-%Y').date()
                queryset = queryset.filter(logdate=date_obj)
            except ValueError:
                # Handle invalid date format
                pass

        # Filter by date range if both date_from and date_to are provided
        if date_from and date_to:
            try:
                date_from_obj = datetime.strptime(date_from, '%m-%d-%Y').date()
                date_to_obj = datetime.strptime(date_to, '%m-%d-%Y').date()
                queryset = queryset.filter(logdate__range=[date_from_obj, date_to_obj]).order_by('logdate')
            except ValueError:
                pass 

        # Filter by non-null late_entry values if late_entry_not_null parameter is provided
        if late_entry_not_null == 'true':
            queryset = queryset.exclude(late_entry__isnull=True).exclude(late_entry='00:00:00')
        if early_exit_not_null == 'true':
            queryset = queryset.exclude(early_exit__isnull=True).exclude(early_exit='00:00:00')
        if overtime_not_null == 'true':
            queryset = queryset.exclude(overtime__isnull=True).exclude(overtime='00:00:00')

        if employee_id:
            queryset = queryset.filter(employeeid=employee_id)

        if company_names:
            company_names_list = [name.strip() for name in company_names.split(',')]
            queryset = queryset.filter(employeeid__company__name__in=company_names_list)
        
        if location_names:
            location_names_list = [name.strip() for name in location_names.split(',')]
            queryset = queryset.filter(employeeid__location__name__in=location_names_list)

        if department_names:
            department_names_list = [name.strip() for name in department_names.split(',')]
            queryset = queryset.filter(employeeid__department__name__in=department_names_list)

        if designation_names:
            designation_names_list = [name.strip() for name in designation_names.split(',')]
            queryset = queryset.filter(employeeid__designation__name__in=designation_names_list)

        # Ensure consistent ordering by a specific field, for example, by 'id'
        # queryset = queryset.order_by('-logdate')

        return queryset

    def get(self, request, *args, **kwargs):
        """
        Get the list of attendance records with optional search query.
        """
        queryset = self.get_queryset()
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

@receiver([post_save, post_delete], sender=Attendance)
def invalidate_and_reload_attendance_cache(sender, instance, **kwargs):
    """
    Signal handler to invalidate and reload the employee cache.
    """
    cache.delete('attendance')
    cache.set('attendance', Attendance.objects.order_by('-logdate').all())

# Pre-load data into cache
cache.get_or_set('attendance', Attendance.objects.order_by('-logdate').all(), timeout=3600)

class ExportAttendanceExcelView(View):
    def get(self, request, *args, **kwargs):

        employee_id = request.GET.get('employee_id')
        employee_name = request.GET.get('employee_name')
        company = request.GET.get('company')
        location = request.GET.get('location')
        department = request.GET.get('department')
        designation = request.GET.get('designation')

        # Get date range parameters
        date_from = self.request.GET.get('date_from') 
        date_to = self.request.GET.get('date_to') 

        queryset = cache.get('attendance')
        if queryset is None:
            print("Fetching queryset from the database.")
            queryset = Attendance.objects.order_by('-logdate').all()
            cache.set('attendance', queryset)
        else:
            print(f"Fetching queryset from the cache. {queryset} ")

        # queryset = Attendance.objects.order_by('-logdate')

        if date_from and date_to:
            try:
                date_from_obj = datetime.strptime(date_from, '%m-%d-%Y').date()
                date_to_obj = datetime.strptime(date_to, '%m-%d-%Y').date()
                queryset = queryset.filter(logdate__range=[date_from_obj, date_to_obj]).order_by('logdate')
            except ValueError:
                pass 

        if employee_id:
            queryset = queryset.filter(Q(employeeid__employee_id__iexact=employee_id))
        if employee_name:
            queryset = queryset.filter(Q(employeeid__employee_name__icontains=employee_name))
        if company:
            queryset = queryset.filter(employeeid__company__name__iexact=company)
        if location:
            queryset = queryset.filter(employeeid__location__name__iexact=location)
        if department:
            queryset = queryset.filter(employeeid__department__name__iexact=department)
        if designation:
            queryset = queryset.filter(employeeid__designation__name__iexact=designation)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        headers = ["Employee ID", "Device Enroll ID", "Employee Name", "Company", "Location", "Job Type", "Department", 
                   "Employee Type", "Desination", "Log Date", "Shift", "Shift Status", "In Time", "Out Time", "Total Hours", 
                   "Late Entry", "Early Exit", "OT Hours"]

        row_num = 1

        # Set font style and background color for headers
        header_font = Font(size=14, bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        footer_font = Font(bold=True)
        footer_fill = PatternFill(start_color="799184", end_color="799184", fill_type="solid")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[ws.cell(row=row_num, column=col_num).column_letter].width = len(header) + 7
        ws.freeze_panes = 'A2'

        for row_num, record in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=record.employeeid.employee_id)
            ws.cell(row=row_num, column=2, value=record.employeeid.device_enroll_id)
            ws.cell(row=row_num, column=3, value=record.employeeid.employee_name)
            ws.cell(row=row_num, column=4, value=record.employeeid.company.name)
            ws.cell(row=row_num, column=5, value=record.employeeid.location.name)
            ws.cell(row=row_num, column=6, value=record.employeeid.job_type)
            ws.cell(row=row_num, column=7, value=record.employeeid.department.name)
            ws.cell(row=row_num, column=8, value=record.employeeid.category)
            ws.cell(row=row_num, column=9, value=record.employeeid.designation.name)
            ws.cell(row=row_num, column=10, value=record.logdate)
            ws.cell(row=row_num, column=11, value=record.employeeid.shift.name)
            ws.cell(row=row_num, column=12, value=record.shift_status)
            ws.cell(row=row_num, column=13, value=record.first_logtime)
            ws.cell(row=row_num, column=14, value=record.last_logtime)
            ws.cell(row=row_num, column=15, value=record.total_time)
            ws.cell(row=row_num, column=16, value=record.late_entry)
            ws.cell(row=row_num, column=17, value=record.early_exit)
            ws.cell(row=row_num, column=18, value=record.overtime)

            # Apply conditional formatting based on Shift Status
            cell = ws.cell(row=row_num, column=12)  # Column 12 is for Shift Status
            if record.shift_status == 'P': 
                # cell.font = header_font
                cell.style = 'Good'

            else: 
                # cell.font = header_font
                cell.style = 'Bad'
                
            cell.alignment = Alignment(horizontal='center')

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=Attendance_Report.xlsx"
        wb.save(response)

        return response 
    
class AttendanceMetricsAPIView(generics.ListAPIView):
    def get_queryset(self):
        # Get the latest log date
        # latest_logdate = datetime.strptime('2023-08-03', '%Y-%m-%d').date()
        latest_logdate = Attendance.objects.latest('logdate').logdate
        # print(latest_logdate)
        # Filter the queryset to include only records with the latest log date
        queryset = Attendance.objects.filter(logdate=latest_logdate)
        return queryset

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        
        # Counting logic
        present_count = queryset.filter(shift_status='P').count()
        absent_count = queryset.filter(shift_status='A').count()
        late_entry_count = queryset.exclude(late_entry__isnull=True).count()
        early_exit_count = queryset.exclude(early_exit__isnull=True).count()
        overtime_count = queryset.exclude(overtime__isnull=True).count()

        latest_logdate = queryset.latest('logdate').logdate

        # counting present count since last hour(current hour - 1 hour)
        present_count_last_hour = queryset.filter(shift_status='P', first_logtime__gte=now()-timedelta(hours=1)).count()

        # Calculate the absent count for the previous week
        previous_week_absent_count = queryset.filter(shift_status='A', logdate__range=[latest_logdate - timedelta(days=7), latest_logdate - timedelta(days=1)]).count()

        # Calculate the absent count for the current week
        current_week_absent_count = queryset.filter(shift_status='A', logdate__range=[latest_logdate - timedelta(days=1), latest_logdate]).count()

        # Calculate the percentage increase in absentees compared to last week
        if previous_week_absent_count != 0:
            absent_percentage_increase = ((current_week_absent_count - previous_week_absent_count) / previous_week_absent_count) * 100
        else:
            absent_percentage_increase = 0

        # Counting employees who arrive late frequently
        late_arrival_threshold = 5  # Define the threshold for frequent late arrivals
        frequent_late_arrivals = queryset.filter(late_entry__isnull=False).annotate(late_count=Count('late_entry')).filter(late_count__gte=late_arrival_threshold).count()

        # Counting employees who arrived late today
        today = date.today()
        late_today = queryset.filter(late_entry__isnull=False, logdate=today).count()
        
        late_arrival_threshold = 5  # Define the threshold for frequent late arrivals
        frequent_late_arrivals = queryset.filter(late_entry__isnull=False).annotate(late_count=Count('late_entry')).filter(late_count__gte=late_arrival_threshold).count()

        current_hour = datetime.now().hour
        # Counting live headcount 
        if 5 <= current_hour <= 17:
            live_headcount = queryset.filter(first_logtime=F('last_logtime')).count()
        else:
            live_headcount = 0    
        # counting total checkin whose first logtime is not null
        total_checkin = queryset.exclude(first_logtime__isnull=True).count()
        # counting total checkout whose last logtime is not null and not equal to first logtime
        total_checkout = queryset.exclude(last_logtime__isnull=True).exclude(last_logtime=F('first_logtime')).count()
        
        # Constructing response data
        data = {
            'present_count': present_count,
            'absent_count': absent_count,
            'late_entry_count': late_entry_count,
            'early_exit_count': early_exit_count,
            'overtime_count': overtime_count,
            'live_headcount': live_headcount,
            'total_checkin': total_checkin,
            'total_checkout': total_checkout,
            'present_count_last_hour': present_count_last_hour,
            'absent_percentage_increase': absent_percentage_increase,
            'frequent_late_arrivals': frequent_late_arrivals,
        }
        
        return Response(data)
    
class AttendanceMonthlyMetricsAPIView(APIView):
    
    def get(self, request, format=None):
        current_date = datetime.now()
        start_date = current_date.replace(day=1)
        
        attendances = Attendance.objects.filter(logdate__range=[start_date, current_date])
        
        # Dictionary to keep track of metrics per day
        metrics_dict = {}

        for attendance in attendances:
            logdate_str = attendance.logdate.strftime('%Y-%m-%d')
            if logdate_str not in metrics_dict:
                metrics_dict[logdate_str] = {
                    'date': logdate_str,
                    'present': 0,
                    'absent': 0,
                    'late_entry': 0,
                    'early_exit': 0,
                    'overtime': 0
                }

            if attendance.shift_status == 'A':
                metrics_dict[logdate_str]['absent'] += 1
            else:
                metrics_dict[logdate_str]['present'] += 1

            if attendance.late_entry and attendance.late_entry != timedelta(0):
                metrics_dict[logdate_str]['late_entry'] += 1
            
            if attendance.early_exit and attendance.early_exit != timedelta(0):
                metrics_dict[logdate_str]['early_exit'] += 1
            
            if attendance.overtime and attendance.overtime != timedelta(0):
                metrics_dict[logdate_str]['overtime'] += 1

        # Convert metrics_dict to a list of daily metrics and sort by date
        daily_metrics = sorted(metrics_dict.values(), key=lambda x: x['date'])

        response_data = {
            'daily_metrics': daily_metrics
        }
        
        return Response(response_data, status=status.HTTP_200_OK)
    
class LogsListCreate(generics.ListCreateAPIView):
    """
    API view for listing and creating logs.
    """
    queryset = Logs.objects.all()
    serializer_class = serializers.LogsSerializer
    pagination_class = DefaultPagination

class LogsRetrieveUpdateDestroy(generics.RetrieveUpdateDestroyAPIView):
    """
    API view for retrieving, updating, and deleting a log.
    """
    queryset = Logs.objects.all()
    serializer_class = serializers.LogsSerializer
    lookup_url_kwarg = "id"